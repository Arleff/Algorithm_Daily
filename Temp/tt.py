# -*- coding: utf-8 -*-
"""
Changed on 20181123

@author: wujiang03
"""

# coding:utf-8
'''
@author caoyu
处理dau转化率by新增bymau，产生月例行的报表，自动化发送邮件

1.初始化excel维表字典
2.读入维表字典的dau、new、mau
3.填表时计算dau转化率
4.填表
5.发送excel表

'''
import copy
import datetime
import getopt
import json
import os
import random
import re
import sys
import time
from collections import defaultdict

import openpyxl as xlrd
from openpyxl import load_workbook
from dau_trans_all import *

dic_year = [31, 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
dic_year_run = [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31]
dic_root = {}  # 基字典
mode_names = [u"excel_mode/201701-201712_DAU转化率_主版_展示表.xlsx", u"excel_mode/201801-201812_DAU转化率_主版_展示表.xlsx",
              u"excel_mode/201701-201712_DAU转化率_lite_展示表.xlsx", u"excel_mode/201801-201812_DAU转化率_lite_展示表.xlsx",
              u"excel_mode/201701-201712_DAU转化率_去重前_主版_展示表.xlsx", u"excel_mode/201801-201812_DAU转化率_去重前_主版_展示表.xlsx",
              u"excel_mode/201701-201712_DAU转化率_去重前_lite_展示表.xlsx",u"excel_mode/201801-201812_DAU转化率_去重前_lite_展示表.xlsx",
              u"excel_mode/201901-201912_DAU转化率_主版_展示表.xlsx",u"excel_mode/201901-201912_DAU转化率_lite_展示表.xlsx"]


def write_excel_1(mode_name, sheet_index, dic, mode):
    # 将转化后数据写入格式化excel中
    data = xlrd.load_workbook(mode_name)
    table = data[sheet_index]
    nrows = table.max_row
    ncolumn = table.max_column
    for i in range(2, nrows + 1):
        new_mon = str(int(table.cell(row=i, column=1).value) * 100 + int(table.cell(row=i, column=2).value[:-1]))
        c1 = table.cell(row=i, column=3).value
        c2 = table.cell(row=i, column=4).value
        c3 = table.cell(row=i, column=5).value
        c4 = table.cell(row=i, column=6).value
        c5 = table.cell(row=i, column=7).value
        j = 12
        for j in range(12, ncolumn):
            if j == 12:
                active_mon = u"0"
            else:
                active_mon = table.cell(row=1, column=j).value[1:]
            # print  new_mon, active_mon, c1, c2, c3, c4, c5
            if c1 == u"Total":
                if dau_trans(dic[new_mon][active_mon], dic[new_mon][u"0"], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = dau_trans(dic[new_mon][active_mon], dic[new_mon][u"0"], mode)
                j += 1
                # print c1
                continue
            if c2 is None or c2 == u"汇总":
                if dau_trans(dic[new_mon][active_mon][c1], dic[new_mon][u"0"][c1], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = dau_trans(dic[new_mon][active_mon][c1], dic[new_mon][u"0"][c1],
                                                              mode)
                j += 1
                continue
            # print c2
            if c3 is None or c3 == u"汇总":
                if dau_trans(dic[new_mon][active_mon][c1][c2], dic[new_mon][u"0"][c1][c2], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = dau_trans(dic[new_mon][active_mon][c1][c2],
                                                              dic[new_mon][u"0"][c1][c2], mode)
                j += 1
                continue
            if c4 is None or c4 == u"汇总":
                if dau_trans(dic[new_mon][active_mon][c1][c2][c3], dic[new_mon][u"0"][c1][c2][c3], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = dau_trans(dic[new_mon][active_mon][c1][c2][c3],
                                                              dic[new_mon][u"0"][c1][c2][c3], mode)
                j += 1
                continue
            if c5 is None or c5 == u"汇总":
                if c4 == u"第三方商店":
                    if c3 == u"IOS":
                        c4 = u"苹果第三方商店"
                    else:
                        c4 = u"安卓第三方商店"
                if dau_trans(dic[new_mon][active_mon][c1][c2][c3][c4], dic[new_mon][u"0"][c1][c2][c3][c4], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = dau_trans(dic[new_mon][active_mon][c1][c2][c3][c4],
                                                              dic[new_mon][u"0"][c1][c2][c3][c4], mode)
                j += 1
                continue
            if dau_trans(dic[new_mon][active_mon][c1][c2][c3][c4][c5], dic[new_mon][u"0"][c1][c2][c3][c4][c5],
                         mode) == 0:
                continue
            table.cell(row=i, column=j).value = dau_trans(dic[new_mon][active_mon][c1][c2][c3][c4][c5],
                                                          dic[new_mon][u"0"][c1][c2][c3][c4][c5], mode)
            j += 1
        i += 1
        # print dic[u"201801"][u"1"][u"Android"][u"预装"][u"厂商"][u"厂商"][u"华为"]
    # print dic[u"201601"][u"1"][u"Android"][u"后装"][u"其他后装"]
    data.save(mode_name)


def write_excel_2(mode_name, sheet_index, dic, mode):
    # 将转化后数据写入格式化excel中
    data = xlrd.load_workbook(mode_name)
    table = data[sheet_index]
    nrows = table.max_row
    ncolumn = table.max_column
    for i in range(2, nrows + 1):
        new_mon = str(int(table.cell(row=i, column=1).value) * 100 + int(table.cell(row=i, column=2).value[:-1]))
        c1 = table.cell(row=i, column=3).value
        c2 = table.cell(row=i, column=4).value
        c3 = table.cell(row=i, column=5).value
        c4 = table.cell(row=i, column=6).value
        c5 = table.cell(row=i, column=7).value
        j = 12
        for j in range(12, ncolumn):
            if j == 12:
                active_mon = u"0"
            else:
                active_mon = table.cell(row=1, column=j).value[1:]
            # print  new_mon, active_mon, c1, c2, c3, c4, c5
            if c1 == u"Total":
                if mau_trans(dic[new_mon][active_mon], dic[new_mon][u"0"], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans(dic[new_mon][active_mon], dic[new_mon][u"0"], mode)
                j += 1
                # print c1
                continue
            if c2 is None or c2 == u"汇总":
                if mau_trans(dic[new_mon][active_mon][c1], dic[new_mon][u"0"][c1], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans(dic[new_mon][active_mon][c1], dic[new_mon][u"0"][c1],
                                                              mode)
                j += 1
                continue
            # print c2
            if c3 is None or c3 == u"汇总":
                if mau_trans(dic[new_mon][active_mon][c1][c2], dic[new_mon][u"0"][c1][c2], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans(dic[new_mon][active_mon][c1][c2],
                                                              dic[new_mon][u"0"][c1][c2], mode)
                j += 1
                continue
            if c4 is None or c4 == u"汇总":
                if mau_trans(dic[new_mon][active_mon][c1][c2][c3], dic[new_mon][u"0"][c1][c2][c3], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans(dic[new_mon][active_mon][c1][c2][c3],
                                                              dic[new_mon][u"0"][c1][c2][c3], mode)
                j += 1
                continue
            if c5 is None or c5 == u"汇总":
                if mau_trans(dic[new_mon][active_mon][c1][c2][c3][c4], dic[new_mon][u"0"][c1][c2][c3][c4], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans(dic[new_mon][active_mon][c1][c2][c3][c4],
                                                              dic[new_mon][u"0"][c1][c2][c3][c4], mode)
                j += 1
                continue
            if mau_trans(dic[new_mon][active_mon][c1][c2][c3][c4][c5], dic[new_mon][u"0"][c1][c2][c3][c4][c5],
                         mode) == 0:
                continue
            table.cell(row=i, column=j).value = mau_trans(dic[new_mon][active_mon][c1][c2][c3][c4][c5],
                                                          dic[new_mon][u"0"][c1][c2][c3][c4][c5], mode)
            j += 1
        i += 1
        # print dic[u"201601"][u"1"][u"Android"][u"后装"][u"其他后装"]
    data.save(mode_name)


def write_excel_3(mode_name, sheet_index, dic, mode):
    # 将转化后数据写入格式化excel中
    data = xlrd.load_workbook(mode_name)
    table = data[sheet_index]
    nrows = table.max_row
    ncolumn = table.max_column
    for i in range(2, nrows + 1):
        new_mon = str(int(table.cell(row=i, column=1).value) * 100 + int(table.cell(row=i, column=2).value[:-1]))
        c1 = table.cell(row=i, column=3).value
        c2 = table.cell(row=i, column=4).value
        c3 = table.cell(row=i, column=5).value
        c4 = table.cell(row=i, column=6).value
        c5 = table.cell(row=i, column=7).value
        j = 12
        for j in range(12, ncolumn):
            if j == 12:
                active_mon = u"0"
            else:
                active_mon = table.cell(row=1, column=j).value[1:]
            # print  new_mon, active_mon, c1, c2, c3, c4, c5
            if c1 == u"Total":
                if mau_trans1(dic[new_mon][active_mon], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans1(dic[new_mon][active_mon], mode)
                j += 1
                # print c1
                continue
            if c2 is None or c2 == u"汇总":
                if mau_trans1(dic[new_mon][active_mon][c1], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans1(dic[new_mon][active_mon][c1], mode)
                j += 1
                continue
            # print c2
            if c3 is None or c3 == u"汇总":
                if mau_trans1(dic[new_mon][active_mon][c1][c2], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans1(dic[new_mon][active_mon][c1][c2], mode)
                j += 1
                continue
            if c4 is None or c4 == u"汇总":
                if mau_trans1(dic[new_mon][active_mon][c1][c2][c3], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans1(dic[new_mon][active_mon][c1][c2][c3], mode)
                j += 1
                continue
            if c5 is None or c5 == u"汇总":
                if mau_trans1(dic[new_mon][active_mon][c1][c2][c3][c4], mode) == 0:
                    continue
                table.cell(row=i, column=j).value = mau_trans1(dic[new_mon][active_mon][c1][c2][c3][c4], mode)
                j += 1
                continue
            if mau_trans1(dic[new_mon][active_mon][c1][c2][c3][c4][c5], mode) == 0:
                continue
            table.cell(row=i, column=j).value = mau_trans1(dic[new_mon][active_mon][c1][c2][c3][c4][c5], mode)
            j += 1
        i += 1
        # print dic[u"201801"][u"1"][u"Android"][u"预装"][u"厂商"][u"厂商"][u"华为"]
    # print dic[u"201601"][u"1"][u"Android"][u"后装"][u"其他后装"]
    data.save(mode_name)


# def send_mails():
#    #将转化后excel表打包发送到oppd组相关数据处理人
#    email_title = "手百分渠道dau转化率月报"
#    email_text = '\n'.join("")
#    email_list = ["caoyu11@baidu.com", "cuijie01@baidu.com"]
#     #    email_list = ["caoyu11@baidu.com",  "wujiang03@baidu.com",  "cuijie01@baidu.com",  "zhaozhixing@baidu.com",  "liukejun01@baidu.com",  "changchenyu@baidu.com"]
#    files = [u"201505-201512_DAU转化率_展示表_0424.xlsx"]
#    os.chdir("/home/work/caoyu/dau_trans/code/")
# #    send_mail.send_job_excute_info(email_title,  email_text,  email_list,  files)

def run(dic_root):
    # 执行程序
    mode_sheets = [u"1.1全量_新增→DAU", u"1.2全量_MAU→DAU", u"1.3付费_新增→DAU", u"1.4付费_MAU→DAU", u"2.1全量_MAU→断代MAU",
                   u"2.2付费_MAU→断代MAU", u"3.1全量_断代MAU→断代DAU", u"3.2付费_断代MAU→断代DAU"]
    caculate_mode = ["new", "mau"]
    data_names = [u"data/data_mainline_all", u"data/data_mainline_pay", u"data/data_lite_all", u"data/data_lite_pay",
                  u"data/data_matrix_mainline_all", u"data/data_matrix_mainline_pay", u"data/data_matrix_lite_all",
                  u"data/data_matrix_lite_pay"]

    # mainline2017
    print
    mode_names[0]
    dic = read_excel_weibiao(mode_names[0], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[0], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[0])
    dic_root_pay1 = read_data(dic_root_pay, data_names[1])
    # print(json.dumps(dic_root1).decode("unicode_escape"))
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[0], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[0], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[0], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[0], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[0], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[0], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[0], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[0], mode_sheets[7], dic_root_pay1, caculate_mode[1])

    # mainline2018
    print
    mode_names[1]
    dic = read_excel_weibiao(mode_names[1], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[1], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[0])
    dic_root_pay1 = read_data(dic_root_pay, data_names[1])
    # print(json.dumps(dic_root1).decode("unicode_escape"))
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[1], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[1], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[1], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[1], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[1], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[1], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[1], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[1], mode_sheets[7], dic_root_pay1, caculate_mode[1])

    # mainline2019
    print
    mode_names[8]
    dic = read_excel_weibiao(mode_names[8], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[8], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[0])
    dic_root_pay1 = read_data(dic_root_pay, data_names[1])
    # print(json.dumps(dic_root1).decode("unicode_escape"))
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[8], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[8], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[8], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[8], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[8], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[8], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[8], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[8], mode_sheets[7], dic_root_pay1, caculate_mode[1])

    # lite2017
    print
    mode_names[2]
    dic = read_excel_weibiao(mode_names[2], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[2], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[2])
    dic_root_pay1 = read_data(dic_root_pay, data_names[3])
    # print(json.dumps(dic_root1).decode("unicode_escape"))
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[2], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[2], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[2], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[2], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[2], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[2], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[2], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[2], mode_sheets[7], dic_root_pay1, caculate_mode[1])

    # lite2018
    print
    mode_names[3]
    dic = read_excel_weibiao(mode_names[3], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[3], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[2])
    dic_root_pay1 = read_data(dic_root_pay, data_names[3])
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[3], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[3], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[3], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[3], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[3], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[3], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[3], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[3], mode_sheets[7], dic_root_pay1, caculate_mode[1])

    # lite2019
    print
    mode_names[9]
    dic = read_excel_weibiao(mode_names[9], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[9], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[2])
    dic_root_pay1 = read_data(dic_root_pay, data_names[3])
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[9], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[9], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[9], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[9], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[9], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[9], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[9], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[9], mode_sheets[7], dic_root_pay1, caculate_mode[1])

    # matrix_mainline2017
    print
    mode_names[4]
    dic = read_excel_weibiao(mode_names[4], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[4], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[4])
    dic_root_pay1 = read_data(dic_root_pay, data_names[5])
    # print(json.dumps(dic_root1).decode("unicode_escape"))
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[4], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[4], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[4], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[4], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[4], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[4], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[4], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[4], mode_sheets[7], dic_root_pay1, caculate_mode[1])

    # matrix_mainline2018
    print
    mode_names[5]
    dic = read_excel_weibiao(mode_names[5], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[5], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[4])
    dic_root_pay1 = read_data(dic_root_pay, data_names[5])
    # print(json.dumps(dic_root1).decode("unicode_escape"))
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[5], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[5], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[5], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[5], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[5], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[5], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[5], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[5], mode_sheets[7], dic_root_pay1, caculate_mode[1])

    # matrix_lite2017
    print
    mode_names[6]
    dic = read_excel_weibiao(mode_names[6], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[6], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[6])
    dic_root_pay1 = read_data(dic_root_pay, data_names[7])
    # print(json.dumps(dic_root1).decode("unicode_escape"))
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[6], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[6], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[6], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[6], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[6], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[6], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[6], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[6], mode_sheets[7], dic_root_pay1, caculate_mode[1])

    # lite2017
    print
    mode_names[7]
    dic = read_excel_weibiao(mode_names[7], mode_sheets[1])
    dic_root = creat_dic(dic, mode_names[7], mode_sheets[1])
    dic_root_pay = copy.deepcopy(dic_root)
    dic_root1 = read_data(dic_root, data_names[6])
    dic_root_pay1 = read_data(dic_root_pay, data_names[7])
    # 变量初始化，跟数据导入

    # 全量
    write_excel_1(mode_names[7], mode_sheets[0], dic_root1, caculate_mode[0])
    write_excel_1(mode_names[7], mode_sheets[1], dic_root1, caculate_mode[1])
    write_excel_2(mode_names[7], mode_sheets[4], dic_root1, caculate_mode[1])
    write_excel_3(mode_names[7], mode_sheets[6], dic_root1, caculate_mode[1])

    # 付费
    write_excel_1(mode_names[7], mode_sheets[2], dic_root_pay1, caculate_mode[0])
    write_excel_1(mode_names[7], mode_sheets[3], dic_root_pay1, caculate_mode[1])
    write_excel_2(mode_names[7], mode_sheets[5], dic_root_pay1, caculate_mode[1])
    write_excel_3(mode_names[7], mode_sheets[7], dic_root_pay1, caculate_mode[1])


if __name__ == "__main__":
    run(dic_root)
